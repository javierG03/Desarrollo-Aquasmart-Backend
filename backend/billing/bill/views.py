from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Sum
from django.utils.timezone import now
from .models import Bill
from .serializers import BillSerializer, BillStatusUpdateSerializer
import io
from .permissions import IsOwnerOrAdmin # Asegúrate de importar tu permiso
from caudal.models import WaterConsumptionRecord
from rest_framework import status
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer


class BillListView(generics.ListAPIView):
    """Vista para obtener todas las facturas o solo las de un usuario."""
    serializer_class = BillSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrAdmin]

    def get_queryset(self):
        """Devuelve solo las facturas del usuario si es un usuario normal, o todas las facturas si es un admin."""
        user = self.request.user
        if user.is_staff:
            return Bill.objects.all() # Administradores pueden ver todas las facturas
        return Bill.objects.filter(client=user) # Usuarios solo pueden ver sus propias facturas
    


class BillDetailView(generics.RetrieveAPIView):
    """Vista para obtener el detalle de una factura específica."""
    queryset = Bill.objects.all()
    serializer_class = BillSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrAdmin] # Asegúrate de que esté el permiso

    def get_object(self):
        obj = super().get_object()
        self.check_object_permissions(self.request, obj) # Verifica permisos
        return obj

# Vista para totalizar facturas con múltiples filtros
class BillTotalizationView(APIView):
    permission_classes = [IsAuthenticated]

    def get_filtered_queryset(self, request):
        status_filter = request.query_params.getlist('status')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        user_id = request.query_params.get('user_id')
        lot_id = request.query_params.get('lot_id')
        plot_name = request.query_params.get('plot_name')

        queryset = Bill.objects.all()

        if status_filter:
            queryset = queryset.filter(status__in=status_filter)

        if start_date and end_date:
            queryset = queryset.filter(creation_date__range=[start_date, end_date])

        if user_id:
            queryset = queryset.filter(client__id=user_id)

        if lot_id:
            queryset = queryset.filter(lot__id=lot_id)

        if plot_name:
            queryset = queryset.filter(plot_name=plot_name)

        return queryset

    def get(self, request):
        queryset = self.get_filtered_queryset(request)

        totalizados = queryset.values('status').annotate(
            cantidad_facturas=Count('id_bill', distinct=True),
            cantidad_usuarios=Count('client', distinct=True),
            cantidad_predios=Count('plot_name', distinct=True),
            cantidad_lotes=Count('lot', distinct=True),
            monto_total=Sum('total_amount')
        )

        return Response(totalizados)

# Exportar totalización a PDF (estilizado)
class ExportTotalizationPDFView(BillTotalizationView):
    def get(self, request):
        queryset = self.get_filtered_queryset(request)

        # Resumen por estado
        resumen = queryset.values('status').annotate(
            cantidad_facturas=Count('id_bill', distinct=True),
            cantidad_usuarios=Count('client', distinct=True),
            cantidad_predios=Count('plot_name', distinct=True),
            cantidad_lotes=Count('lot', distinct=True),
            monto_total=Sum('total_amount')
        )

        # Tabla detallada
        detalles = queryset.values_list(
            'code', 'status', 'client_name', 'plot_name', 'lot_code', 'creation_date', 'total_amount'
        )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Título
        elements.append(Paragraph("<b>Informe de Facturas</b>", styles['Title']))
        elements.append(Spacer(1, 12))

        # Tabla resumen
        resumen_data = [['Estado', 'Facturas', 'Usuarios', 'Predios', 'Lotes', 'Monto Total']]
        for row in resumen:
            resumen_data.append([
                row['status'].capitalize(),
                row['cantidad_facturas'],
                row['cantidad_usuarios'],
                row['cantidad_predios'],
                row['cantidad_lotes'],
                f"${float(row['monto_total'] or 0):,.2f}"
            ])
        resumen_table = Table(resumen_data, hAlign='LEFT')
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(resumen_table)
        elements.append(Spacer(1, 24))

        # Tabla de facturas
        elements.append(Paragraph("<b>Facturas Detalladas</b>", styles['Heading2']))
        detalle_data = [['Código', 'Estado', 'Cliente', 'Predio', 'Lote', 'Fecha', 'Total']]
        for d in detalles:
            detalle_data.append([
                d[0], d[1].capitalize(), d[2], d[3], d[4],
                d[5].strftime('%Y-%m-%d'), f"${float(d[6] or 0):,.2f}"
            ])
        detalle_table = Table(detalle_data, hAlign='LEFT')
        detalle_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#444444')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(detalle_table)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="informe_facturas_{now().strftime("%Y%m%d")}.pdf"'
        return response



# Exportar totalización a Excel (estilizado)
class ExportTotalizationExcelView(BillTotalizationView):
    def get(self, request):
        queryset = self.get_filtered_queryset(request)

        resumen = queryset.values('status').annotate(
            cantidad_facturas=Count('id_bill', distinct=True),
            cantidad_usuarios=Count('client', distinct=True),
            cantidad_predios=Count('plot_name', distinct=True),
            cantidad_lotes=Count('lot', distinct=True),
            monto_total=Sum('total_amount')
        )

        detalles = queryset.values_list(
            'code', 'status', 'client_name', 'plot_name', 'lot_code', 'creation_date', 'total_amount'
        )

        wb = openpyxl.Workbook()
        ws_resumen = wb.active
        ws_resumen.title = "Resumen de Facturas"

        headers = ['Estado', 'Facturas', 'Usuarios', 'Predios', 'Lotes', 'Monto Total']
        ws_resumen.append(headers)

        for col in ws_resumen[1]:
            col.font = Font(bold=True, color="FFFFFF")
            col.fill = PatternFill("solid", fgColor="003366")
            col.alignment = Alignment(horizontal="center")

        for r in resumen:
            ws_resumen.append([
                r['status'].capitalize(),
                r['cantidad_facturas'],
                r['cantidad_usuarios'],
                r['cantidad_predios'],
                r['cantidad_lotes'],
                float(r['monto_total'] or 0)
            ])

        # Segunda hoja: Detalles
        ws_detalle = wb.create_sheet("Facturas Detalladas")
        detalle_headers = ['Código', 'Estado', 'Cliente', 'Predio', 'Lote', 'Fecha', 'Total']
        ws_detalle.append(detalle_headers)

        for col in ws_detalle[1]:
            col.font = Font(bold=True, color="FFFFFF")
            col.fill = PatternFill("solid", fgColor="444444")
            col.alignment = Alignment(horizontal="center")

        for d in detalles:
            ws_detalle.append([
                d[0], d[1].capitalize(), d[2], d[3], d[4], d[5].strftime('%Y-%m-%d'), float(d[6] or 0)
            ])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="informe_facturas_{now().strftime("%Y%m%d")}.xlsx"'
        return response



from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.conf import settings
from .models import Bill
from plots_lots.models import Lot
from billing.models import Company
import hmac
import hashlib
import time
from datetime import datetime, timedelta

@csrf_exempt
@require_POST
def generate_monthly_bills(request):
    """Endpoint para generar facturas mensuales mediante una solicitud HTTP"""
    # Verificación de seguridad con token secreto
    provided_token = request.headers.get('X-Api-Token')

    if not settings.BILLING_SECRET_TOKEN or not provided_token:
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    # Comparación segura de tokens
    if not hmac.compare_digest(settings.BILLING_SECRET_TOKEN, provided_token):
        return JsonResponse({'error': 'Invalid token'}, status=403)
    
    # Obtener la empresa que emite las facturas
    company = Company.objects.first()
    if not company:
        return JsonResponse({'error': 'No company found'}, status=404)

    active_lots = Lot.objects.filter(is_activate=True)
    bills_created = 0
    errors = []

    # Fecha actual
    today = datetime.now().date()
    first_day = today.replace(day=1)
    last_day = (first_day + timedelta(days=32)).replace(day=1) - timedelta(days=1)

    for lot in active_lots:
        try:
            client = lot.plot.owner

            # Verificar si ya existe una factura para este lote en el mes actual
            existing_bill = Bill.objects.filter(
                lot=lot,
                creation_date__gte=first_day,
                creation_date__lte=last_day
            ).exists()

            if existing_bill:
                continue

            # Verificar si el cliente está inactivo
            is_client_inactive = not getattr(client, 'is_active', True)

            # Buscar la última factura para este cliente
            if is_client_inactive:
                last_client_bill = Bill.objects.filter(client=client).order_by('-creation_date').first()
                if last_client_bill and getattr(last_client_bill, 'client_inactive', False):
                    # Ya se le facturó una vez estando inactivo, no crear nueva factura
                    continue

            last_consumption = WaterConsumptionRecord.objects.filter(
                lot=lot,
                billed=False,
                end_date__lt=today # Solo consumos con periodos ya cerrados
            ).order_by('-end_date').first()

            volumetric_quantity = last_consumption.period_consumption if last_consumption else getattr(lot, 'estimated_consumption', 2)

            # Crear factura
            bill = Bill(
                company=company,
                client=client,
                lot=lot,
                fixed_rate_quantity=1,
                volumetric_rate_quantity=volumetric_quantity
            )

            # Si el cliente está inactivo, marcar
            if is_client_inactive and hasattr(Bill, 'client_inactive'):
                bill.client_inactive = True

            if last_consumption:
                last_consumption.billed = True

            bill.save()
            if last_consumption:
                last_consumption.save()
            bills_created += 1

        except Exception as e:
            errors.append(f"Error al procesar lote {lot.id_lot}: {str(e)}")

    return JsonResponse({
        'status': 'success',
        'bills_created': bills_created,
        'date': today.strftime('%Y-%m-%d'),
        'errors': errors
    })    


class UpdateBillStatusAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = BillStatusUpdateSerializer(data=request.data)
        if serializer.is_valid():
            code = serializer.validated_data['code']
            try:
                bill = Bill.objects.get(code=code)
            except Bill.DoesNotExist:
                return Response({"detail": "Factura no encontrada."}, status=status.HTTP_404_NOT_FOUND)

            if bill.client != request.user:
                return Response({"detail": "No tienes permiso para modificar esta factura."}, status=status.HTTP_403_FORBIDDEN)

            if bill.status == 'pagada':
                return Response({"detail": "La factura ya está marcada como pagada."}, status=status.HTTP_400_BAD_REQUEST)

            bill.status = 'pagada'
            bill.save()
            return Response({"detail": f" Pago exitoso de la factura {bill.code}."}, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)    
